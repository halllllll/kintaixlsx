# テスト

import os
import settings
import styles
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl import Workbook

def main():
    now = datetime.now()
    next_month = now + relativedelta(months=1)
    pre_month = now - relativedelta(months=1)
    print(next_month, pre_month)
    date_suffix = "{0}{1:02d}".format(now.year, now.month)
    files = os.listdir(os.getcwd())
    file_name = "{}_{}.xlsx".format(settings.XL_BASE_NAME, date_suffix)

    wb = Workbook()

    if file_name not in files:
        ws = wb.active
        ws.title = date_suffix
        ws["A1"].font = styles.A1
        ws["A1"].value = "{0}月分勤務表（{1}年{2}月16日～{3}年{4}月15日）".format(now.month, pre_month.year, pre_month.month, next_month.year, now.month)
        ws["A2"].font = styles.A2
        ws["A2"].value = settings.A2_DEFAULT
        ws.merge_cells("A2:F2")
        ws["G2"].font = styles.G2
        ws["G2"].alignment = styles.G2_Align
        ws["G2"].value = settings.EMPLOYEE_NAME
        ws.column_dimensions["G"].width = 40

    else:
        wb = load_workbook(file_name)

    ws = wb.active
    wb.save(file_name)
if __name__ == "__main__":
    main()