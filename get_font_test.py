# セルの情報とってこれる？
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = r"Z:\101個人\06川島\work_schedule\monthly\勤務表.xlsx")
ws = wb.active
a1 = ws["A1"]
print(a1)
print(a1.value)
print(a1.font)
print(a1.font.name, a1.font.size)
print(a1.font.color.rgb)
# print(a1.style)
# print(dir(a1))
# print(a1.fill)
